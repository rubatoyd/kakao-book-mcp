"""Exporters and safe_name tests."""
from __future__ import annotations

import csv
import json
import sqlite3
from pathlib import Path

from kakao_book_mcp.exporters import export, safe_name
from kakao_book_mcp.models import BookRecord


def test_safe_name():
    assert safe_name("normal_name") == "normal_name"
    assert safe_name("../../etc/passwd") == "etc_passwd"
    assert safe_name("CON.txt") == "kakao_book_output"
    assert safe_name("파이썬 프로그래밍: 기초?") == "파이썬_프로그래밍__기초"


def test_export_all_formats(tmp_path: Path):
    records = [
        BookRecord(
            title="테스트 도서 1",
            authors="홍길동",
            publisher="테스트출판사",
            pub_year="2023",
            pub_date="2023-01-01",
            price=20000,
            sale_price=18000,
            discount_rate=10.0,
            isbn="1234567890 9781234567890",
            isbn10="1234567890",
            isbn13="9781234567890",
            status="정상판매",
            contents="내용 요약 1",
        ),
        BookRecord(
            title="테스트 도서 2",
            authors="이순신",
            publisher="해양출판사",
            pub_year="2024",
            pub_date="2024-02-01",
            price=15000,
            sale_price=15000,
            discount_rate=None,
            isbn="9789876543210",
            isbn10="",
            isbn13="9789876543210",
            status="정상판매",
            contents="내용 요약 2",
        ),
    ]

    formats = ["json", "csv", "xlsx", "sqlite"]
    created = export(records, formats, tmp_path, "sample_books")
    assert len(created) == 4

    # 1. JSON 검증
    json_path = tmp_path / "sample_books.json"
    assert json_path.exists()
    data = json.loads(json_path.read_text(encoding="utf-8"))
    assert len(data) == 2
    assert data[0]["title"] == "테스트 도서 1"

    # 2. CSV 검증
    csv_path = tmp_path / "sample_books.csv"
    assert csv_path.exists()
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = list(csv.DictReader(f))
        assert len(reader) == 2
        assert reader[0]["title"] == "테스트 도서 1"
        assert reader[0]["authors"] == "홍길동"

    # 3. XLSX 검증
    xlsx_path = tmp_path / "sample_books.xlsx"
    assert xlsx_path.exists()
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.title == "books"
    assert ws.cell(row=2, column=2).value == "테스트 도서 1"

    # 4. SQLite 검증
    sqlite_path = tmp_path / "sample_books.sqlite"
    assert sqlite_path.exists()
    conn = sqlite3.connect(sqlite_path)
    cur = conn.cursor()
    cur.execute("SELECT title, authors, price FROM books")
    rows = cur.fetchall()
    conn.close()
    assert len(rows) == 2
    assert rows[0][0] == "테스트 도서 1"
    assert rows[0][1] == "홍길동"
    assert rows[0][2] == "20000"
