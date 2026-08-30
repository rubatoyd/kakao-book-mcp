"""수집 결과(BookRecord)를 xlsx/csv/json/sqlite 로 저장."""
from __future__ import annotations

import csv
import json
import re
import sqlite3
from pathlib import Path
from typing import Sequence

from .models import COLUMNS, BookRecord

# Windows 예약 장치명 방어
_RESERVED = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}
_UNSAFE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def safe_name(name: str, *, fallback: str = "kakao_book_output", limit: int = 60) -> str:
    """디렉터리 이탈 및 위험 문자를 방지하는 안전한 파일명 정규화."""
    s = _UNSAFE.sub("_", str(name or ""))
    s = s.replace("..", "_").strip().strip(". ")
    s = re.sub(r"\s+", "_", s)[:limit].strip("._ ")
    if not s or s.upper().split(".")[0] in _RESERVED:
        s = fallback
    return s


def _rows(records: Sequence[BookRecord]) -> list[dict]:
    return [r.to_row() for r in records]


def to_json(records: Sequence[BookRecord], path: str | Path) -> None:
    """정규화 행 및 원본 API 필드(raw)를 함께 JSON으로 저장."""
    data = [{**r.to_row(), "raw": r.raw} for r in records]
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def to_csv(records: Sequence[BookRecord], path: str | Path) -> None:
    """Excel 한글 호환을 위해 UTF-8 BOM(utf-8-sig)으로 CSV 저장."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(_rows(records))


def to_xlsx(records: Sequence[BookRecord], path: str | Path) -> None:
    """openpyxl을 사용하여 Excel 통합 문서로 저장."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "books"

    # 헤더 작성 및 서식 적용
    ws.append(COLUMNS)
    header_fill = PatternFill(start_color="334E68", end_color="334E68", fill_type="solid")
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="Malgun Gothic", size=10)
    for r in records:
        row_dict = r.to_row()
        row_values = [row_dict.get(c, "") for c in COLUMNS]
        ws.append(row_values)

    # 데이터 서식
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)

    wb.save(path)


def to_sqlite(records: Sequence[BookRecord], path: str | Path, table_name: str = "books") -> None:
    """SQLite 데이터베이스 파일로 저장."""
    p = Path(path)
    conn = sqlite3.connect(p)
    cursor = conn.cursor()

    col_defs = ", ".join(f'"{c}" TEXT' for c in COLUMNS)
    cursor.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({col_defs})')
    cursor.execute(f'DELETE FROM "{table_name}"')

    placeholders = ", ".join("?" for _ in COLUMNS)
    rows_data = []
    for r in records:
        row_dict = r.to_row()
        rows_data.append([str(row_dict.get(c, "")) for c in COLUMNS])

    cursor.executemany(
        f'INSERT INTO "{table_name}" VALUES ({placeholders})',
        rows_data
    )
    conn.commit()
    conn.close()


def export(
    records: Sequence[BookRecord],
    formats: Sequence[str],
    out_dir: str | Path,
    base_name: str,
) -> list[str]:
    """요청된 모든 포맷으로 파일 저장 후 생성된 파일 경로 목록 반환."""
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    safe_base = safe_name(base_name)

    created: list[str] = []
    fmt_set = {f.lower().strip().lstrip(".") for f in formats}

    if "json" in fmt_set:
        fpath = out_path / f"{safe_base}.json"
        to_json(records, fpath)
        created.append(str(fpath))

    if "csv" in fmt_set:
        fpath = out_path / f"{safe_base}.csv"
        to_csv(records, fpath)
        created.append(str(fpath))

    if "xlsx" in fmt_set:
        fpath = out_path / f"{safe_base}.xlsx"
        to_xlsx(records, fpath)
        created.append(str(fpath))

    if "sqlite" in fmt_set or "db" in fmt_set:
        fpath = out_path / f"{safe_base}.sqlite"
        to_sqlite(records, fpath)
        created.append(str(fpath))

    return created
